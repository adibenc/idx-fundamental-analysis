import openpyxl

from builders.fundamental_analyser import FundamentalAnalyser
from utils.logger_config import logger


class Excel:
    def __init__(self, filename: str, fundamental_analyser: FundamentalAnalyser):
        self.filename = filename
        self.fundamental_analyser = fundamental_analyser

        try:
            # Try to load an existing workbook
            self.wb = openpyxl.load_workbook(self.filename)
        except FileNotFoundError:
            # Create a new workbook if it doesn't exist
            self.wb = openpyxl.Workbook()

            # Remove the default sheet created.
            self.wb.remove(self.wb.active)

    def _write_to_sheet(self, sheet_name: str, values: []):
        """
        Write values to an existing or new sheet in the Excel file.

        :param sheet_name: Name of the sheet
        :param values: List of rows (each row is a list of values)
        """
        # Create a new sheet if it doesn't exist
        if sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(title=sheet_name)

        sheet = self.wb[sheet_name]

        for i, row_data in enumerate(values):
            for j, value in enumerate(row_data):
                sheet.cell(row=i + 1, column=j + 1, value=value)

    def save(self):
        """
        Save the workbook to a file.
        """
        self.wb.save(self.filename)
        logger.info(
            f"Excel file saved successfully in the root project (./{self.filename})"
        )

    def insert_stock(self):
        """
        Inserts stock data into the spreadsheet.
        """
        self._write_to_sheet("idx-stocks", self.fundamental_analyser.stocks_sheet())

    def insert_key_statistic(self):
        """
        Inserts keystatistic data into the spreadsheet.
        """
        self._write_to_sheet(
            "key-statistics", self.fundamental_analyser.key_statistics_sheet()
        )

    def insert_analysis(self):
        """
        Inserts fundamental analysis data into the spreadsheet.
        """

        self._write_to_sheet("analysis", self.fundamental_analyser.analysis_sheet())
